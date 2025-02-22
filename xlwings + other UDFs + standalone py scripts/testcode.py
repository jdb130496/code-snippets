import sqlite3
from openpyxl import load_workbook
from datetime import datetime

# Define the file path and sheet name
file_path = 'Z:\\Private\\dhawal\\TransactionDetailWithProjectID-2024.xlsm'
sheet_name = 'TransactionDetailWithProj'

wb = load_workbook(filename=file_path, data_only=True)
ws = wb[sheet_name]

# Create a connection to an in-memory SQLite database
conn = sqlite3.connect(':memory:')
cursor = conn.cursor()

# Create the table
cursor.execute('''
CREATE TABLE a (
    ProjectSOPO TEXT,
    Class_Name TEXT,
    Type TEXT,
    Date TEXT,
    Document_Number TEXT,
    Name TEXT,
    Memo TEXT,
    Account TEXT,
    Clr TEXT,
    Split TEXT,
    Qty REAL,
    Amount REAL
)
''')

# Function to convert Excel date to string format
def convert_date(excel_date):
if isinstance(excel_date, datetime):
return excel_date.strftime('%Y-%m-%d')
return excel_date

# Read the data from the specified range in the Excel file and insert into the SQLite database
for row in ws.iter_rows(min_row=6, max_row=43962, min_col=1, max_col=12, values_only=True):
row = list(row)
row[3] = convert_date(row[3])  # Convert the Date column to string format
cursor.execute('''
INSERT INTO a (ProjectSOPO, Class_Name, Type, Date, Document_Number, Name, Memo, Account, Clr, Split, Qty, Amount)
VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', row)

# Commit the transaction
conn.commit()

# Define the SQL query
query = """
SELECT IFNULL(NULLIF(ProjectSOPO, ''), 'Unknown') AS ProjectSOPO, Class_Name, SUM(Amount) AS Amt
FROM a
WHERE (Type = 'Invoice' AND Date BETWEEN '2024-01-01' AND '2024-12-31' AND Class_Name != 'Carvart')
GROUP BY IFNULL(NULLIF(ProjectSOPO, ''), 'Unknown'), Class_Name;
"""

# Execute the SQL query
result = cursor.execute(query).fetchall()

# Print the result
print("\nSQL Query Result:")
for row in result:
print(row)

# Check if the troubled project is included
troubled_project = [row for row in result if row[0] == '34402 - 7 Platt Street']
print("\nTroubled Project in SQL Query Result:")
print(troubled_project)

# Close the connection
conn.close()
