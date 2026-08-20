import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import re
import xlwings as xw
import os

# ─────────────────────────────────────────────
# STEP 1: Load the source file
# ─────────────────────────────────────────────
SOURCE_FILE = r"D:\RR Data\RR AR Investments.xlsm"
SHEET_NAME  = "Fixed Deposits Banks"
OUTPUT_FILE = r"D:\RR Data\FD_Quarterly_Schedule.xlsx"

# Read the raw sheet - no header yet, read everything as strings

# ─────────────────────────────────────────────
# READ VIA XLWINGS - works even if file is open in Excel
# ─────────────────────────────────────────────
import shutil, tempfile, math

# Copy file to temp location — works whether Excel has it open or not
temp_dir  = tempfile.mkdtemp()
temp_file = os.path.join(temp_dir, os.path.basename(SOURCE_FILE))
shutil.copy2(SOURCE_FILE, temp_file)
print(f"Reading from temp copy: {temp_file}")

raw = pd.read_excel(temp_file, sheet_name=SHEET_NAME, header=None, dtype=str)
raw = raw.fillna("nan")
print(f"Read {raw.shape[0]} rows x {raw.shape[1]} cols")

# Clean up temp file
shutil.rmtree(temp_dir, ignore_errors=True)

# ── DIAGNOSTIC: print first 5 rows to see actual content and column positions ──
print("\n--- RAW FIRST 5 ROWS ---")
print(raw.head(5).to_string())
print("\n--- RAW COLUMNS 5 TO 10 OF ROWS 2-10 (Int Rate area) ---")
print(raw.iloc[1:10, 5:10].to_string())
print("--- END DIAGNOSTIC ---\n")
# ─────────────────────────────────────────────
# STEP 2: Find the header row
# The header row contains "INVESTMENT DETAILS" or similar in column 0
# ─────────────────────────────────────────────
header_row_idx = None
for i, row in raw.iterrows():
    col0 = str(row[0]).strip().lower()
    col1 = str(row[1]).strip().lower()
    col8 = str(row[8]).strip().lower()
    # True header row has "investment" in col 0, "date" in col 1, "int rate" in col 8
    if "investment" in col0 and col1 == "date" and col8 == "int rate":
        header_row_idx = i
        break

if header_row_idx is None:
    raise ValueError("Could not find header row - expected 'investment' in col 0, 'date' in col 1, 'int rate' in col 8")

print(f"Header row found at index: {header_row_idx}")
print(f"Header row content col 0-10: {list(raw.iloc[header_row_idx, 0:10])}")

# ─────────────────────────────────────────────
# STEP 3: Find the stop row - "Total" appears in Int Rate column (col index 7)
# Stop at that row (do not include it or anything after)
# ─────────────────────────────────────────────
#INT_RATE_COL = 7   # 0-based column index for "Int Rate"
INT_RATE_COL = 8   # 0-based: col 7 is "Cumulative/Non-Cumulative", col 8 is "Int Rate"

stop_row_idx = None
for i in range(header_row_idx + 1, len(raw)):
    val = str(raw.iloc[i, INT_RATE_COL]).strip().lower()
    if val == "total":
        stop_row_idx = i
        break

if stop_row_idx is None:
    stop_row_idx = len(raw)   # no total line found - use all rows

print(f"Stop row (Total line) at index: {stop_row_idx}")

# ─────────────────────────────────────────────
# STEP 4: Slice data rows (between header and stop)
# ─────────────────────────────────────────────
data_raw = raw.iloc[header_row_idx + 1 : stop_row_idx].copy()
data_raw.columns = range(data_raw.shape[1])
print(f"Data slice: {len(data_raw)} rows before cleaning")

# ─────────────────────────────────────────────
# STEP 5: Assign column names based on known positions
# Col 0=Name, 1=Date, 2=IntDates, 3=Maturity, 4=Folio, 5=Holders,
# 6=Term, 7=IntRate, 8=Principal, 9=Nominee, 10=QtrlyInt, 11=TDS, 12=NetInt
# ─────────────────────────────────────────────
#data_raw.columns = [
#    "Name", "StartDate", "IntDates", "Maturity", "Folio",
#    "Holders", "Term", "IntRate", "Principal", "Nominee",
#    "QtrlyAmt", "TDS", "NetInt"
#] + [f"Extra{i}" for i in range(data_raw.shape[1] - 13)]
data_raw.columns = [
    "Name", "StartDate", "IntDates", "Maturity", "Folio",
    "Holders", "Term", "CumNonCum", "IntRate", "Principal", "Nominee",
    "QtrlyAmt", "TDS", "NetInt"
] + [f"Extra{i}" for i in range(data_raw.shape[1] - 14)]


# ─────────────────────────────────────────────
# STEP 6: Drop completely blank rows
# A row is blank if Name, StartDate, Maturity, and Principal are all empty/NaN
# ─────────────────────────────────────────────
def is_blank(val):
    return pd.isna(val) or str(val).strip() in ("", "nan", "None")

mask_blank = (
    data_raw["Name"].apply(is_blank) &
    data_raw["StartDate"].apply(is_blank) &
    data_raw["Maturity"].apply(is_blank) &
    data_raw["Principal"].apply(is_blank)
)
data_raw = data_raw[~mask_blank].copy()
data_raw.reset_index(drop=True, inplace=True)
print(f"After dropping blank rows: {len(data_raw)} rows")

# ─────────────────────────────────────────────
# STEP 7: Drop rows that are PPF / non-FD entries
# PPF rows have no Int Rate (they are not FDs and have no quarterly payout)
# Also drop rows where Principal is blank
# ─────────────────────────────────────────────
mask_no_rate = data_raw["IntRate"].apply(is_blank)
mask_no_principal = data_raw["Principal"].apply(is_blank)
dropped = data_raw[mask_no_rate | mask_no_principal]
if len(dropped):
    print(f"Dropping {len(dropped)} rows with no interest rate or no principal (e.g. PPF):")
    for _, r in dropped.iterrows():
        print(f"  -> {r['Name']} | Date: {r['StartDate']}")

data_raw = data_raw[~(mask_no_rate | mask_no_principal)].copy()
data_raw.reset_index(drop=True, inplace=True)
print(f"After dropping non-FD rows: {len(data_raw)} rows")

# ─────────────────────────────────────────────
# STEP 8: Clean and parse individual fields
# ─────────────────────────────────────────────

# 8a: Clean Name - strip whitespace
data_raw["Name"] = data_raw["Name"].apply(lambda x: str(x).strip() if not is_blank(x) else "")

# 8b: Parse dates - handle both string dates and Excel serial numbers inline
# First try parsing as date string, then fall back to Excel serial number
data_raw["StartDate"] = pd.to_datetime(data_raw["StartDate"], dayfirst=True, errors="coerce").dt.date
data_raw["StartDate"] = pd.Series([
    (pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(float(v)))).date()
    if pd.isnull(r) and str(v).strip().lstrip("-").isdigit()
    else r
    for r, v in zip(data_raw["StartDate"], data_raw["StartDate"].astype(str))
])

data_raw["Maturity"] = pd.to_datetime(data_raw["Maturity"], dayfirst=True, errors="coerce").dt.date
data_raw["Maturity"] = pd.Series([
    (pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(float(v)))).date()
    if pd.isnull(r) and str(v).strip().lstrip("-").isdigit()
    else r
    for r, v in zip(data_raw["Maturity"], data_raw["Maturity"].astype(str))
])

# 8c: Clean Principal - remove commas, spaces, currency symbols, convert to float
def parse_amount(val):
    if is_blank(val):
        return None
    val = str(val).strip().replace(",", "").replace(" ", "").replace("₹", "").replace("$", "")
    try:
        return float(val)
    except:
        return None

data_raw["Principal"] = data_raw["Principal"].apply(parse_amount)

# 8d: Clean IntRate - convert to float (e.g. "6.50" -> 6.50)
def parse_rate(val):
    if is_blank(val):
        return None
    val = str(val).strip().replace("%", "")
    try:
        return float(val)
    except:
        return None

data_raw["IntRate"] = data_raw["IntRate"].apply(parse_rate)

# 8e: Clean Folio - strip quotes and extra whitespace
data_raw["Folio"] = data_raw["Folio"].apply(
    lambda x: str(x).strip().replace('"', "").replace("\n", " | ") if not is_blank(x) else ""
)

# 8f: Clean Holders and Nominee
data_raw["Holders"] = data_raw["Holders"].apply(lambda x: str(x).strip() if not is_blank(x) else "")
data_raw["Nominee"] = data_raw["Nominee"].apply(lambda x: str(x).strip() if not is_blank(x) else "")

# 8g: Clean Term
data_raw["Term"] = data_raw["Term"].apply(lambda x: str(x).strip() if not is_blank(x) else "")

print("\nSample of parsed data:")
print(data_raw[["Name", "StartDate", "Maturity", "Principal", "IntRate"]].to_string())

# ─────────────────────────────────────────────
# STEP 9: Sort - primary: StartDate, secondary: Name (bank name)
# ─────────────────────────────────────────────
data_raw.sort_values(by=["StartDate", "Name"], inplace=True, na_position="last")
data_raw.reset_index(drop=True, inplace=True)
print(f"\nSorted {len(data_raw)} FD records by date then bank name")

# ─────────────────────────────────────────────
# STEP 10: Generate quarterly interest schedule for each FD
# ─────────────────────────────────────────────
from dateutil.relativedelta import relativedelta

output_rows = []

for idx, fd in data_raw.iterrows():
    name       = fd["Name"]
    holders    = fd["Holders"]
    nominee    = fd["Nominee"]
    folio      = fd["Folio"]
    start_date = fd["StartDate"]
    maturity   = fd["Maturity"]
    term       = fd["Term"]
    rate       = fd["IntRate"]
    principal  = fd["Principal"]

    # Skip if essential data missing - explicit None/NaT check
    if start_date is None or maturity is None or principal is None or rate is None:
        print(f"  Skipping row {idx} '{name}' - missing essential data")
        continue
    if pd.isnull(start_date) or pd.isnull(maturity):
        print(f"  Skipping row {idx} '{name}' - unparseable date")
        continue
    # First interest date is 3 months after start date
    current_date = start_date + relativedelta(months=3)
    quarter_num  = 1

    while current_date <= maturity:

        # Start of this interest period (end of previous quarter)
        period_start = start_date + relativedelta(months=3 * (quarter_num - 1))

        # What would the next full quarter date be
        next_full_quarter = start_date + relativedelta(months=3 * quarter_num)

        # Is this a broken/short last period?
        if current_date == maturity and next_full_quarter > maturity:
            # Last period is shorter than a full quarter - use actual days
            actual_days     = (maturity - period_start).days
            period_interest = principal * (rate / 100) * (actual_days / 365)
            print(f"  Broken last period for '{name}': {period_start} to {maturity} "
                  f"= {actual_days} days, interest = {period_interest:.2f}")
        else:
            # Full quarter - standard 3/12
            period_interest = principal * (rate / 100) * (3 / 12)

        tds_amount   = period_interest * 0.10
        net_interest = period_interest - tds_amount

        is_maturity = (current_date == maturity)

        if is_maturity:
            interest_on_date  = period_interest
            tds_on_date       = tds_amount
            net_on_date       = net_interest
            principal_return  = principal
            total_receipt     = net_on_date + principal_return
        else:
            interest_on_date  = period_interest
            tds_on_date       = tds_amount
            net_on_date       = net_interest
            principal_return  = 0
            total_receipt     = net_on_date

        output_rows.append({
            "Bank / FD Name"  : name,
            "Holders"         : holders,
            "Nominee"         : nominee,
            "Folio"           : folio,
            "Start Date"      : start_date,
            "Maturity Date"   : maturity,
            "Term"            : term,
            "Rate %"          : rate,
            "Principal"       : principal,
            "Interest Date"   : current_date,
            "Quarter #"       : quarter_num,
            "Qtrly Interest"  : round(interest_on_date, 2),
            "TDS (10%)"       : round(tds_on_date, 2),
            "Net Interest"    : round(net_on_date, 2),
            "Principal Return": round(principal_return, 2),
            "Total Receipt"   : round(total_receipt, 2),
            "Is Maturity"     : is_maturity,
        })

        # Stop after maturity row
        if is_maturity:
            break

        # Advance to next quarter
        quarter_num += 1
        next_date = start_date + relativedelta(months=3 * quarter_num)

        # If next full quarter overshoots maturity, jump to maturity instead
        if next_date > maturity:
            current_date = maturity
        else:
            current_date = next_date

print(f"\nTotal output rows (all quarters for all FDs): {len(output_rows)}")

# ─────────────────────────────────────────────
# STEP 11: Build output DataFrame
# ─────────────────────────────────────────────
out_df = pd.DataFrame(output_rows)

# Sort all quarterly rows by Interest Date (primary) then Bank Name (secondary)
out_df.sort_values(by=["Interest Date", "Bank / FD Name"], ascending=[True, True], inplace=True)
out_df.reset_index(drop=True, inplace=True)

print(f"Output dataframe shape: {out_df.shape}")
print(out_df.head(10).to_string())

# ─────────────────────────────────────────────
# STEP 12: Write to Excel using openpyxl with formatting
# ─────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "FD Quarterly Schedule"

# Define styles
header_font  = Font(name="Arial", bold=True, size=10)
header_fill  = PatternFill()   # no fill
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font    = Font(name="Arial", size=9)
maturity_fill = PatternFill()  # no fill
alt_fill     = PatternFill()   # no fill

thin = Side(style="thin", color="000000")
border = Border(bottom=thin)   # only bottom border, no side borders

date_format    = "DD-MMM-YYYY"
amount_format  = '#,##0.00'
currency_format= '₹#,##0.00'

# Column headers and widths
columns = [
    ("Bank / FD Name",  30),
    ("Holders",         18),
    ("Nominee",         14),
    ("Folio",           22),
    ("Start Date",      14),
    ("Maturity Date",   14),
    ("Term",            12),
    ("Rate %",          8),
    ("Principal",       16),
    ("Interest Date",   14),
    ("Qtrly Interest",  16),
    ("TDS (10%)",       14),
    ("Net Interest",    14),
    ("Principal Return",16),
    ("Total Receipt",   16),
]

# Write header row
for col_idx, (col_name, col_width) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font    = header_font
    cell.fill    = header_fill
    cell.alignment = header_align
    cell.border  = border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 30

# Map column names to output_df columns
col_map = {
    "Bank / FD Name"  : "Bank / FD Name",
    "Holders"         : "Holders",
    "Nominee"         : "Nominee",
    "Folio"           : "Folio",
    "Start Date"      : "Start Date",
    "Maturity Date"   : "Maturity Date",
    "Term"            : "Term",
    "Rate %"          : "Rate %",
    "Principal"       : "Principal",
    "Interest Date"   : "Interest Date",
    "Qtrly Interest"  : "Qtrly Interest",
    "TDS (10%)"       : "TDS (10%)",
    "Net Interest"    : "Net Interest",
    "Principal Return": "Principal Return",
    "Total Receipt"   : "Total Receipt",
}

# Track FD group for alternating color (by Start Date + Name combo)
prev_fd_key = None
use_alt_fill = False

# Write data rows
for row_idx, record in enumerate(out_df.to_dict("records"), start=2):    
    fd_key = (record["Bank / FD Name"], record["Start Date"], record["Folio"])
    if fd_key != prev_fd_key:
        use_alt_fill = not use_alt_fill
        prev_fd_key = fd_key
    is_mat = record["Is Maturity"]
    row_fill = maturity_fill if is_mat else (alt_fill if use_alt_fill else PatternFill())
    for col_idx, (col_name, _) in enumerate(columns, start=1):
        df_key = col_map[col_name]
        value  = record.get(df_key, "")
        cell   = ws.cell(row=row_idx, column=col_idx)

        # Write value - dates as Excel serial integers for correct sorting
        if col_name in ("Start Date", "Maturity Date", "Interest Date"):
            if value is not None and hasattr(value, 'toordinal'):
                cell.value = (value - date(1899, 12, 30)).days
            else:
                cell.value = value
            cell.number_format = date_format
            cell.alignment = Alignment(horizontal="center")
        elif col_name in ("Principal", "Qtrly Interest", "TDS (10%)",
                          "Net Interest", "Principal Return", "Total Receipt"):
            cell.value = value
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal="right")
        elif col_name == "Rate %":
            cell.value = value
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.value = value
            cell.alignment = Alignment(horizontal="left", wrap_text=False)

        cell.font   = data_font
        cell.border = border
        cell.fill   = row_fill
        
    # Bold the maturity row for visual emphasis
    #if is_mat:
    #    for col_idx in range(1, len(columns) + 1):
    #        ws.cell(row=row_idx, column=col_idx).font = Font(name="Arial", size=9, bold=True)

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter on header row
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

# ─────────────────────────────────────────────
# STEP 13: Add a summary sheet - one row per FD
# ─────────────────────────────────────────────
ws2 = wb.create_sheet(title="FD Summary")

summary_cols = [
    ("Bank / FD Name",  30),
    ("Holders",         18),
    ("Nominee",         14),
    ("Folio",           22),
    ("Start Date",      14),
    ("Maturity Date",   14),
    ("Term",            12),
    ("Rate %",          8),
    ("Principal",       16),
    ("Annual Interest", 16),
    ("Annual TDS",      14),
    ("Annual Net Int",  14),
]

for col_idx, (col_name, col_width) in enumerate(summary_cols, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=col_name)
    cell.font    = header_font
    cell.fill    = header_fill
    cell.alignment = header_align
    cell.border  = border
    ws2.column_dimensions[get_column_letter(col_idx)].width = col_width

ws2.row_dimensions[1].height = 30

summary_row = 2
for idx, fd in data_raw.iterrows():
    principal  = fd["Principal"]
    rate       = fd["IntRate"]

    # Skip if rate or principal is None (should not happen after cleaning, but guard anyway)
    if principal is None or rate is None:
        print(f"  Summary: skipping '{fd['Name']}' - principal or rate is None")
        continue

    annual_int = principal * (rate / 100)
    annual_tds = annual_int * 0.10
    annual_net = annual_int - annual_tds

    summary_data = [
        fd["Name"],
        fd["Holders"],
        fd["Nominee"],
        fd["Folio"],
        fd["StartDate"],
        fd["Maturity"],
        fd["Term"],
        rate,
        principal,
        round(annual_int, 2),
        round(annual_tds, 2),
        round(annual_net, 2),
    ]

    for col_idx, value in enumerate(summary_data, start=1):
        cell = ws2.cell(row=summary_row, column=col_idx, value=value)
        cell.font   = data_font
        cell.border = border
        col_name = summary_cols[col_idx - 1][0]
        if col_name in ("Start Date", "Maturity Date"):
            cell.number_format = date_format
            cell.alignment = Alignment(horizontal="center")
        elif col_name in ("Principal", "Annual Interest", "Annual TDS", "Annual Net Int"):
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal="right")
        elif col_name == "Rate %":
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")
    summary_row += 1

# Total row on summary sheet
total_label_col = 1
total_row = summary_row
ws2.cell(row=total_row, column=total_label_col, value="TOTAL").font = Font(name="Arial", bold=True, size=9)

for col_idx, (col_name, _) in enumerate(summary_cols, start=1):
    if col_name in ("Principal", "Annual Interest", "Annual TDS", "Annual Net Int"):
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell = ws2.cell(row=total_row, column=col_idx, value=formula)
        cell.font = Font(name="Arial", bold=True, size=9)
        cell.number_format = currency_format
        cell.fill = PatternFill("solid", fgColor="D9EAD3")  # light green
        cell.border = border
        cell.alignment = Alignment(horizontal="right")

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(summary_cols))}1"

# ─────────────────────────────────────────────
# STEP 14: Save the workbook
# ─────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\n✅ Output saved to: {OUTPUT_FILE}")
print(f"   Sheet 1 'FD Quarterly Schedule': {len(output_rows)} rows")
print(f"   Sheet 2 'FD Summary': {len(data_raw)} FDs")
