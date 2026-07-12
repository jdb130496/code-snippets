import re
import sys
import traceback
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

INPUT_FILE = "painting.xml"
OUTPUT_FILE = "painting.xlsx"

try:
    with open(INPUT_FILE, "r", encoding="utf-16") as f:
        content = f.read()

    TAGS = ["DSPVCHDATE", "DSPVCHLEDACCOUNT", "NAMEFIELD", "INFOFIELD",
            "DSPVCHTYPE", "DSPVCHDRAMT", "DSPVCHCRAMT",
            "VCHLEDNARREXPLOSION", "DSPVCHNARR"]

    tag_alt = "|".join(TAGS)
    pattern = re.compile(rf"<({tag_alt})>(.*?)</\1>", re.DOTALL)
    matches = pattern.findall(content)
    print(f"Total tags found: {len(matches)}", flush=True)

    def clean(text):
        text = text.strip()
        text = text.replace("&amp;", "&").replace("&quot;", '"')
        text = text.replace("&lt;", "<").replace("&gt;", ">")
        text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
        return text

    def parse_date(value):
        """Convert 'D-M-YYYY' or 'DD-MM-YYYY' text to a real date object."""
        try:
            return datetime.strptime(value, "%d-%m-%Y").date()
        except ValueError:
            return value  # fall back to raw text if format doesn't match

    def parse_amount(value):
        """Convert numeric text to float, keep blank as empty string."""
        if not value:
            return ""
        try:
            return float(value)
        except ValueError:
            return value

    records = []
    current = {}

    for tag, value in matches:
        value = clean(value)

        if tag == "DSPVCHDATE" and value:
            if current:
                records.append(current)
            current = {"Date": parse_date(value)}
            continue

        if tag == "DSPVCHLEDACCOUNT" and value:
            current["Ledger Account"] = value
        elif tag == "DSPVCHTYPE" and value:
            current["Voucher Type"] = value
        elif tag == "DSPVCHDRAMT" and value:
            current["Debit Amount"] = parse_amount(value)
        elif tag == "DSPVCHCRAMT" and value:
            current["Credit Amount"] = parse_amount(value)
        elif tag == "VCHLEDNARREXPLOSION" and value:
            current["Narration (Explosion)"] = value
        elif tag == "DSPVCHNARR" and value:
            current["Narration"] = value

    if current:
        records.append(current)

    print(f"Total records parsed: {len(records)}", flush=True)

    if not records:
        print("WARNING: No records parsed - check tag structure.", flush=True)
        sys.exit(1)

    columns = ["Date", "Ledger Account", "Voucher Type", "Debit Amount",
               "Credit Amount", "Narration (Explosion)", "Narration"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Painting Vouchers"

    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, record in enumerate(records, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = record.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_name == "Date" and isinstance(value, __import__("datetime").date):
                cell.number_format = "DD-MM-YYYY"
            elif col_name in ("Debit Amount", "Credit Amount") and isinstance(value, float):
                cell.number_format = "#,##0.00"

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            [len(str(record.get(col_name, ""))) for record in records] + [len(col_name)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(OUTPUT_FILE)
    print(f"Saved {len(records)} records to {OUTPUT_FILE}", flush=True)

except Exception:
    print("\n--- ERROR TRACEBACK ---", flush=True)
    traceback.print_exc()
    input("\nPress Enter to close...")
